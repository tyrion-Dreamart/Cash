from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Optional
import models
from database import get_db
from currency import to_usd, DEFAULT_FX_RATE as FX

router = APIRouter(prefix="/export", tags=["Export"])

def fill(hex): return PatternFill("solid", start_color=hex, fgColor=hex)
def bdr():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

DARK="111827"; GREEN_D="065F46"; RED_D="7F1D1D"; BLUE_D="1E3A5F"
GREEN_L="D1FAE5"; RED_L="FEE2E2"; YELLOW_L="FEF3C7"; GRAY="F9FAFB"

def make_excel(cxc_rows, cxp_rows, others_rows=[], fx=FX):
    wb = Workbook()
    today = date.today()

    # ── RESUMEN ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [25,18,18,18,14,16]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    ws["A1"] = "DREAMART PHOTOGRAPHY GROUP — CXC vs CXP"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = fill(DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generado: {today.strftime('%d/%m/%Y')}  |  FX: {fx} MXN/USD"
    ws["A2"].font = Font(name="Arial", size=9, color="6B7280")
    ws["A2"].fill = fill(GRAY); ws["A2"].alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16
    ws.append([])

    row = 4
    for col, hdr in enumerate(["País","CXC — Nos deben","CXP — Debemos","Neto","Ratio","Estado"],1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = fill(DARK); c.border = bdr()
        c.alignment = Alignment(horizontal="center" if col>1 else "left", indent=1 if col==1 else 0)
    ws.row_dimensions[row].height = 20

    countries = {}
    for r in cxc_rows:
        c = r.country or "Sin país"
        bal = to_usd(float(r.amount) - float(r.amount_paid or 0), r.currency, fx)
        countries.setdefault(c, {"cxc":0,"cxp":0})
        countries[c]["cxc"] += bal
    for r in cxp_rows:
        c = r.country or "Sin país"
        bal = to_usd(float(r.amount) - float(r.amount_paid or 0), r.currency, fx)
        countries.setdefault(c, {"cxc":0,"cxp":0})
        countries[c]["cxp"] += bal

    total_cxc = total_cxp = 0
    for country in sorted(countries.keys()):
        d = countries[country]
        net = d["cxc"] - d["cxp"]
        ratio = round(d["cxc"]/d["cxp"],2) if d["cxp"]>0 else 0
        row += 1
        bg = GRAY if row%2==0 else "FFFFFF"
        c1 = ws.cell(row=row, column=1, value=country)
        c1.font = Font(name="Arial", bold=True, size=10); c1.fill=fill(bg); c1.border=bdr(); c1.alignment=Alignment(indent=1)
        for col, val, color, bg2 in [(2,d["cxc"],GREEN_D,GREEN_L),(3,d["cxp"],RED_D,RED_L),(4,net,GREEN_D if net>=0 else RED_D,GREEN_L if net>=0 else RED_L)]:
            c = ws.cell(row=row, column=col, value=round(val,2))
            c.font=Font(name="Arial",bold=True,size=10,color=color); c.fill=fill(bg2); c.border=bdr(); c.number_format='$#,##0.00'; c.alignment=Alignment(horizontal="right")
        ws.cell(row=row,column=5,value=ratio).font=Font(name="Arial",size=10,color=GREEN_D if ratio>=1 else RED_D)
        ws.cell(row=row,column=5).fill=fill(bg); ws.cell(row=row,column=5).border=bdr(); ws.cell(row=row,column=5).number_format='0.00x'; ws.cell(row=row,column=5).alignment=Alignment(horizontal="center")
        ws.cell(row=row,column=6,value="Positivo" if net>=0 else "Deficit").font=Font(name="Arial",size=10,color=GREEN_D if net>=0 else RED_D)
        ws.cell(row=row,column=6).fill=fill(bg); ws.cell(row=row,column=6).border=bdr(); ws.cell(row=row,column=6).alignment=Alignment(horizontal="center")
        ws.row_dimensions[row].height=18; total_cxc+=d["cxc"]; total_cxp+=d["cxp"]

    net_total = total_cxc - total_cxp
    row += 1
    ws.row_dimensions[row].height = 22
    for col, val, color, bg in [(1,"TOTAL","FFFFFF",DARK),(2,round(total_cxc,2),"FFFFFF",GREEN_D),(3,round(total_cxp,2),"FFFFFF",RED_D),(4,round(net_total,2),"FFFFFF",GREEN_D if net_total>=0 else RED_D),(5,round(total_cxc/total_cxp,2) if total_cxp>0 else 0,"FFFFFF",DARK),(6,"OK" if net_total>=0 else "Deficit","FFFFFF",DARK)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font=Font(name="Arial",bold=True,size=11,color=color); c.fill=fill(bg); c.border=bdr()
        if col in [2,3,4]: c.number_format='$#,##0.00'; c.alignment=Alignment(horizontal="right")
        elif col==5: c.number_format='0.00x'; c.alignment=Alignment(horizontal="center")
        elif col==1: c.alignment=Alignment(horizontal="left",indent=1)
        else: c.alignment=Alignment(horizontal="center")

    # ── CXC POR CLIENTE ────────────────────────────────────
    ws2 = wb.create_sheet("CXC - Por Cliente")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJKLM", [28,14,14,14,14,14,14,10,12,12,20,16,35]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "CXC — CUENTAS POR COBRAR (Agrupado por cliente)"
    ws2["A1"].font=Font(name="Arial",bold=True,size=12,color="FFFFFF"); ws2["A1"].fill=fill(GREEN_D)
    ws2["A1"].alignment=Alignment(horizontal="left",vertical="center",indent=1); ws2.row_dimensions[1].height=26

    ws2.merge_cells("A2:H2")
    ws2["A2"] = f"Generado: {today.strftime('%d/%m/%Y')}  |  FX: {fx} MXN/USD  |  Rojo = Vencida"
    ws2["A2"].font=Font(name="Arial",size=9,color="6B7280"); ws2["A2"].fill=fill(GRAY); ws2["A2"].alignment=Alignment(horizontal="left",indent=1)

    row = 3
    for col, h in enumerate(["Cliente","# Factura","Fecha Factura","Monto USD","Monto MXN","Saldo USD","Saldo MXN","Moneda","Vto.","Status","Hotel","País","Notas"],1):
        c = ws2.cell(row=row,column=col,value=h)
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(GREEN_D); c.border=bdr(); c.alignment=Alignment(horizontal="center")
    ws2.row_dimensions[row].height=20

    # Group by client
    from collections import defaultdict
    cxc_by_client = defaultdict(list)
    for r in cxc_rows:
        cxc_by_client[r.client_name].append(r)

    cxc_total = 0
    for client in sorted(cxc_by_client.keys()):
        records = cxc_by_client[client]
        client_total = sum(to_usd(float(r.amount)-float(r.amount_paid or 0), r.currency, fx) for r in records)
        cxc_total += client_total

        # Client header
        row += 1
        ws2.merge_cells(f"A{row}:J{row}")
        c = ws2.cell(row=row,column=1,value=f"  {client.upper()}")
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(BLUE_D); c.alignment=Alignment(horizontal="left",vertical="center")
        ws2.row_dimensions[row].height=18
        c2 = ws2.cell(row=row,column=11,value=f"Total: ${client_total:,.2f} USD")
        c2.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c2.fill=fill(BLUE_D); c2.alignment=Alignment(horizontal="right")
        for _c in [12,13]: ws2.cell(row=row,column=_c).fill=fill(BLUE_D)

        for i, r in enumerate(records):
            row += 1
            due = r.due_date
            is_overdue = due and due < today
            orig = to_usd(float(r.amount), r.currency, fx)
            bal = to_usd(float(r.amount)-float(r.amount_paid or 0), r.currency, fx)
            bg = RED_L if is_overdue else (GRAY if i%2==0 else "FFFFFF")
            import re as _re
            factura = ""
            notas = r.comments or ""
            if r.comments:
                m = _re.search(r"F/(\S+)", r.comments)
                if m: factura = "F/" + m.group(1)
                notas = _re.sub(r"Factura[:\s]*F?\s*/?\s*[\w]+", "", r.comments)
                notas = _re.sub(r"Pais:\s*[\wÀ-ɏ\s]+", "", notas)
                notas = _re.sub(r"F\s*/\s*[\w]+", "", notas)
                notas = _re.sub(r"^[Ff]actura\s*$", "", notas.strip())
                notas = notas.strip(" |,").strip()
                if notas.lower() in ["factura", "factura:", ""]: notas = ""
            country_val = r.country or "SIN PAIS"
            cur = str(r.currency).replace("Currency.","")
            orig_mxn = round(float(r.amount) * fx, 2) if cur == "USD" else round(float(r.amount), 2)
            bal_mxn = round((float(r.amount) - float(r.amount_paid or 0)) * fx, 2) if cur == "USD" else round(float(r.amount) - float(r.amount_paid or 0), 2)
            data = [r.client_name, factura, str(r.invoice_date) if r.invoice_date else "--", round(orig,2), round(orig_mxn,2) if cur != "MXN" else "--", round(bal,2), round(bal_mxn,2) if cur != "MXN" else "--", cur, str(due) if due else "--", str(r.status).replace("CXCStatus.",""), r.hotel or "--", country_val, notas]
            for col, val in enumerate(data,1):
                c = ws2.cell(row=row,column=col,value=val)
                c.border=bdr(); c.fill=fill(bg)
                if col in [2,3]: c.number_format='$#,##0.00'; c.alignment=Alignment(horizontal="right"); c.font=Font(name="Arial",size=9,color=RED_D if is_overdue else GREEN_D)
                elif col==5: c.font=Font(name="Arial",size=9,bold=is_overdue,color=RED_D if is_overdue else "111827"); c.alignment=Alignment(horizontal="center")
                elif col==6: c.font=Font(name="Arial",size=9,color="111827"); c.alignment=Alignment(horizontal="center")
                elif col==8: c.font=Font(name="Arial",size=8,color="4B5563"); c.alignment=Alignment(wrap_text=True)
                else: c.font=Font(name="Arial",size=9,color="111827"); c.alignment=Alignment(horizontal="left")
            ws2.row_dimensions[row].height=15

    row += 1
    ws2.merge_cells(f"A{row}:B{row}")
    c=ws2.cell(row=row,column=1,value="TOTAL CXC"); c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=fill(GREEN_D); c.border=bdr(); c.alignment=Alignment(indent=1)
    c=ws2.cell(row=row,column=3,value=round(cxc_total,2)); c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=fill(GREEN_D); c.border=bdr(); c.number_format='$#,##0.00'; c.alignment=Alignment(horizontal="right")
    for col in [4,5,6,7,8,9,10,11,12,13]: ws2.cell(row=row,column=col).fill=fill(GREEN_D); ws2.cell(row=row,column=col).border=bdr()
    ws2.row_dimensions[row].height=24

    # ── CXP POR PROVEEDOR ──────────────────────────────────
    ws3 = wb.create_sheet("CXP - Por Proveedor")
    ws3.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJKLMN", [28,14,14,14,14,14,14,10,12,12,10,20,16,35]):
        ws3.column_dimensions[col].width = w

    ws3.merge_cells("A1:I1")
    ws3["A1"] = "CXP — CUENTAS POR PAGAR (Agrupado por proveedor)"
    ws3["A1"].font=Font(name="Arial",bold=True,size=12,color="FFFFFF"); ws3["A1"].fill=fill(RED_D)
    ws3["A1"].alignment=Alignment(horizontal="left",vertical="center",indent=1); ws3.row_dimensions[1].height=26

    ws3.merge_cells("A2:I2")
    ws3["A2"] = f"Generado: {today.strftime('%d/%m/%Y')}  |  FX: {fx} MXN/USD  |  Rojo=Vencida, Amarillo=Alta prioridad"
    ws3["A2"].font=Font(name="Arial",size=9,color="6B7280"); ws3["A2"].fill=fill(GRAY); ws3["A2"].alignment=Alignment(horizontal="left",indent=1)

    row = 3
    for col, h in enumerate(["Proveedor","# Factura","Fecha Factura","Monto USD","Monto MXN","Saldo USD","Saldo MXN","Moneda","Vto.","Status","Prioridad","Hotel","País","Notas"],1):
        c = ws3.cell(row=row,column=col,value=h)
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(RED_D); c.border=bdr(); c.alignment=Alignment(horizontal="center")
    ws3.row_dimensions[row].height=20

    cxp_by_vendor = defaultdict(list)
    for r in cxp_rows:
        cxp_by_vendor[r.vendor_name].append(r)

    cxp_total = 0
    for vendor in sorted(cxp_by_vendor.keys()):
        records = cxp_by_vendor[vendor]
        vendor_total = sum(to_usd(float(r.amount)-float(r.amount_paid or 0), r.currency, fx) for r in records)
        cxp_total += vendor_total

        row += 1
        ws3.merge_cells(f"A{row}:K{row}")
        c = ws3.cell(row=row,column=1,value=f"  {vendor.upper()}")
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(BLUE_D); c.alignment=Alignment(horizontal="left",vertical="center")
        ws3.row_dimensions[row].height=18
        c2 = ws3.cell(row=row,column=12,value=f"Total: ${vendor_total:,.2f} USD")
        c2.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c2.fill=fill(BLUE_D); c2.alignment=Alignment(horizontal="right")
        for _c in [13,14]: ws3.cell(row=row,column=_c).fill=fill(BLUE_D)

        for i, r in enumerate(records):
            row += 1
            due = r.due_date
            is_overdue = due and due < today
            priority = str(r.priority).replace("CXPPriority.","") if r.priority else "media"
            is_alta = priority == "alta"
            orig = to_usd(float(r.amount), r.currency, fx)
            bal = to_usd(float(r.amount)-float(r.amount_paid or 0), r.currency, fx)
            if is_overdue and is_alta: bg=RED_L
            elif is_alta: bg=YELLOW_L
            elif i%2==0: bg=GRAY
            else: bg="FFFFFF"
            import re as _re2
            factura2 = ""
            notas2 = r.comments or ""
            if r.comments:
                m2 = _re2.search(r"F/(\S+)", r.comments)
                if m2: factura2 = "F/" + m2.group(1)
                notas2 = _re2.sub(r"Factura[:\s]*F?\s*/?\s*[\w]+", "", r.comments)
                notas2 = _re2.sub(r"Pais:\s*[\wÀ-ɏ\s]+", "", notas2)
                notas2 = _re2.sub(r"F\s*/\s*[\w]+", "", notas2)
                notas2 = notas2.strip(" |,").strip()
                if notas2.lower() in ["factura", "factura:", ""]: notas2 = ""
            country_val2 = r.country or "SIN PAIS"
            cur2 = str(r.currency).replace("Currency.","")
            orig_mxn2 = round(float(r.amount) * fx, 2) if cur2 == "USD" else round(float(r.amount), 2)
            bal_mxn2 = round((float(r.amount) - float(r.amount_paid or 0)) * fx, 2) if cur2 == "USD" else round(float(r.amount) - float(r.amount_paid or 0), 2)
            data = [r.vendor_name, factura2, str(r.invoice_date) if r.invoice_date else "--", round(orig,2), round(orig_mxn2,2) if cur2 != "MXN" else "--", round(bal,2), round(bal_mxn2,2) if cur2 != "MXN" else "--", cur2, str(due) if due else "--", str(r.status).replace("CXPStatus.",""), priority, r.hotel or "--", country_val2, notas2]
            for col, val in enumerate(data,1):
                c = ws3.cell(row=row,column=col,value=val)
                c.border=bdr(); c.fill=fill(bg)
                if col in [2,3]: c.number_format='$#,##0.00'; c.alignment=Alignment(horizontal="right"); c.font=Font(name="Arial",size=9,color=RED_D if (is_overdue or is_alta) else "111827")
                elif col in [5,6,7]: c.font=Font(name="Arial",size=9,color=RED_D if is_alta else "111827"); c.alignment=Alignment(horizontal="center")
                elif col==9: c.font=Font(name="Arial",size=8,color="4B5563"); c.alignment=Alignment(wrap_text=True)
                else: c.font=Font(name="Arial",size=9,color="111827"); c.alignment=Alignment(horizontal="left")
            ws3.row_dimensions[row].height=15

    row += 1
    ws3.merge_cells(f"A{row}:B{row}")
    c=ws3.cell(row=row,column=1,value="TOTAL CXP"); c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=fill(RED_D); c.border=bdr(); c.alignment=Alignment(indent=1)
    c=ws3.cell(row=row,column=3,value=round(cxp_total,2)); c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=fill(RED_D); c.border=bdr(); c.number_format='$#,##0.00'; c.alignment=Alignment(horizontal="right")
    for col in [4,5,6,7,8,9,10,11,12,13,14]: ws3.cell(row=row,column=col).fill=fill(RED_D); ws3.cell(row=row,column=col).border=bdr()
    ws3.row_dimensions[row].height=24

    # ── OTHERS ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Otros - Others")
    ws4.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [30,10,15,15,15,13,13,20,35]):
        ws4.column_dimensions[col].width = w

    ws4.merge_cells("A1:I1")
    ws4["A1"] = "OTROS — Reembolsos / Anticipos / Garantias"
    ws4["A1"].font=Font(name="Arial",bold=True,size=12,color="FFFFFF"); ws4["A1"].fill=fill("7C3AED")
    ws4["A1"].alignment=Alignment(horizontal="left",vertical="center",indent=1); ws4.row_dimensions[1].height=26

    ws4.merge_cells("A2:I2")
    ws4["A2"] = f"Generado: {date.today().strftime('%d/%m/%Y')}  |  FX: {fx} MXN/USD"
    ws4["A2"].font=Font(name="Arial",size=9,color="6B7280"); ws4["A2"].fill=fill(GRAY); ws4["A2"].alignment=Alignment(horizontal="left",indent=1)

    row = 3
    for col, h in enumerate(["Concepto","Direccion","Monto USD","Monto MXN","Saldo USD","Moneda","Vencimiento","Contraparte","Notas"],1):
        c = ws4.cell(row=row,column=col,value=h)
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill("7C3AED"); c.border=bdr(); c.alignment=Alignment(horizontal="center")
    ws4.row_dimensions[row].height=20

    others_total_pay = 0
    others_total_collect = 0
    for i, o in enumerate(others_rows):
        row += 1
        due = o.due_date
        is_overdue = due and due < date.today()
        direction = str(o.direction).replace("OtherDirection.","")
        orig = to_usd(float(o.amount), o.currency, fx)
        bal = to_usd(float(o.amount) - float(o.amount_paid or 0), o.currency, fx)
        cur = str(o.currency).replace("Currency.","")
        orig_mxn = round(float(o.amount) * fx, 2) if cur == "USD" else round(float(o.amount), 2)

        if direction == "pagar": others_total_pay += bal
        else: others_total_collect += bal

        bg = RED_L if (is_overdue and direction=="pagar") else (GREEN_L if direction=="cobrar" else (GRAY if i%2==0 else "FFFFFF"))
        data = [o.concept or "--", direction, round(orig,2), round(orig_mxn,2) if cur!="MXN" else "--", round(bal,2), cur, str(due) if due else "--", o.counterparty or "--", o.comments or "--"]
        for col, val in enumerate(data,1):
            c = ws4.cell(row=row,column=col,value=val)
            c.border=bdr(); c.fill=fill(bg)
            if col in [3,4,5]: c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right"); c.font=Font(name="Arial",size=9,color=RED_D if (is_overdue and direction=="pagar") else GREEN_D if direction=="cobrar" else "111827")
            elif col==7: c.font=Font(name="Arial",size=9,bold=is_overdue,color=RED_D if is_overdue else "111827"); c.alignment=Alignment(horizontal="center")
            else: c.font=Font(name="Arial",size=9,color="111827"); c.alignment=Alignment(horizontal="left")
        ws4.row_dimensions[row].height=15

    # Others totals
    row += 1
    ws4.merge_cells(f"A{row}:B{row}")
    c=ws4.cell(row=row,column=1,value="TOTAL A PAGAR"); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(RED_D); c.border=bdr(); c.alignment=Alignment(indent=1)
    c=ws4.cell(row=row,column=5,value=round(others_total_pay,2)); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(RED_D); c.border=bdr(); c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right")
    for col in [3,4,6,7,8,9]: ws4.cell(row=row,column=col).fill=fill(RED_D); ws4.cell(row=row,column=col).border=bdr()
    ws4.row_dimensions[row].height=20

    row += 1
    ws4.merge_cells(f"A{row}:B{row}")
    c=ws4.cell(row=row,column=1,value="TOTAL A COBRAR"); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(GREEN_D); c.border=bdr(); c.alignment=Alignment(indent=1)
    c=ws4.cell(row=row,column=5,value=round(others_total_collect,2)); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(GREEN_D); c.border=bdr(); c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right")
    for col in [3,4,6,7,8,9]: ws4.cell(row=row,column=col).fill=fill(GREEN_D); ws4.cell(row=row,column=col).border=bdr()
    ws4.row_dimensions[row].height=20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_others_excel(others_rows, fx=FX):
    wb = Workbook()
    today = date.today()

    ws = wb.active
    ws.title = "Otros - Others"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [30,10,15,15,15,13,13,20,35]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    ws["A1"] = "DREAMART PHOTOGRAPHY GROUP — OTROS (Reembolsos / Anticipos / Garantias)"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = fill("7C3AED")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generado: {today.strftime('%d/%m/%Y')}  |  FX: {fx} MXN/USD"
    ws["A2"].font = Font(name="Arial", size=9, color="6B7280")
    ws["A2"].fill = fill(GRAY); ws["A2"].alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16

    row = 3
    for col, h in enumerate(["Concepto","Direccion","Monto USD","Monto MXN","Saldo USD","Moneda","Vencimiento","Contraparte","Notas"],1):
        c = ws.cell(row=row,column=col,value=h)
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill("7C3AED"); c.border=bdr(); c.alignment=Alignment(horizontal="center")
    ws.row_dimensions[row].height=20

    others_total_pay = 0
    others_total_collect = 0
    for i, o in enumerate(others_rows):
        row += 1
        due = o.due_date
        is_overdue = due and due < today
        direction = str(o.direction).replace("OtherDirection.","")
        orig = to_usd(float(o.amount), o.currency, fx)
        bal = to_usd(float(o.amount) - float(o.amount_paid or 0), o.currency, fx)
        cur = str(o.currency).replace("Currency.","")
        orig_mxn = round(float(o.amount) * fx, 2) if cur == "USD" else round(float(o.amount), 2)

        if direction == "pagar": others_total_pay += bal
        else: others_total_collect += bal

        bg = RED_L if (is_overdue and direction=="pagar") else (GREEN_L if direction=="cobrar" else (GRAY if i%2==0 else "FFFFFF"))
        data = [o.concept or "--", direction, round(orig,2), round(orig_mxn,2) if cur!="MXN" else "--", round(bal,2), cur, str(due) if due else "--", o.counterparty or "--", o.comments or "--"]
        for col, val in enumerate(data,1):
            c = ws.cell(row=row,column=col,value=val)
            c.border=bdr(); c.fill=fill(bg)
            if col in [3,4,5]: c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right"); c.font=Font(name="Arial",size=9,color=RED_D if (is_overdue and direction=="pagar") else GREEN_D if direction=="cobrar" else "111827")
            elif col==7: c.font=Font(name="Arial",size=9,bold=is_overdue,color=RED_D if is_overdue else "111827"); c.alignment=Alignment(horizontal="center")
            else: c.font=Font(name="Arial",size=9,color="111827"); c.alignment=Alignment(horizontal="left")
        ws.row_dimensions[row].height=15

    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    c=ws.cell(row=row,column=1,value="TOTAL A PAGAR"); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(RED_D); c.border=bdr(); c.alignment=Alignment(indent=1)
    c=ws.cell(row=row,column=5,value=round(others_total_pay,2)); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(RED_D); c.border=bdr(); c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right")
    for col in [3,4,6,7,8,9]: ws.cell(row=row,column=col).fill=fill(RED_D); ws.cell(row=row,column=col).border=bdr()
    ws.row_dimensions[row].height=20

    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    c=ws.cell(row=row,column=1,value="TOTAL A COBRAR"); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(GREEN_D); c.border=bdr(); c.alignment=Alignment(indent=1)
    c=ws.cell(row=row,column=5,value=round(others_total_collect,2)); c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF"); c.fill=fill(GREEN_D); c.border=bdr(); c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right")
    for col in [3,4,6,7,8,9]: ws.cell(row=row,column=col).fill=fill(GREEN_D); ws.cell(row=row,column=col).border=bdr()
    ws.row_dimensions[row].height=20

    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    net = others_total_collect - others_total_pay
    c=ws.cell(row=row,column=1,value="NETO"); c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=fill(DARK); c.border=bdr(); c.alignment=Alignment(indent=1)
    c=ws.cell(row=row,column=5,value=round(net,2)); c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=fill(DARK); c.border=bdr(); c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right")
    for col in [3,4,6,7,8,9]: ws.cell(row=row,column=col).fill=fill(DARK); ws.cell(row=row,column=col).border=bdr()
    ws.row_dimensions[row].height=22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/cxc-cxp")
def export_cxc_cxp(
    fx_rate: float = Query(default=17.5),
    country: Optional[str] = None,
    hotel: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status_cxc: Optional[str] = None,
    status_cxp: Optional[str] = None,
    db: Session = Depends(get_db)
):
    # CXC
    q_cxc = db.query(models.Receivable).filter(models.Receivable.status != "cobrado")
    if country: q_cxc = q_cxc.filter(models.Receivable.country == country)
    if hotel: q_cxc = q_cxc.filter(models.Receivable.hotel.ilike(f"%{hotel}%"))
    if date_from: q_cxc = q_cxc.filter(models.Receivable.due_date >= date_from)
    if date_to: q_cxc = q_cxc.filter(models.Receivable.due_date <= date_to)
    if status_cxc: q_cxc = q_cxc.filter(models.Receivable.status == status_cxc)
    cxc_rows = q_cxc.order_by(models.Receivable.client_name, models.Receivable.due_date).all()

    # CXP
    q_cxp = db.query(models.Payable).filter(models.Payable.status != "pagado")
    if country: q_cxp = q_cxp.filter(models.Payable.country == country)
    if hotel: q_cxp = q_cxp.filter(models.Payable.hotel.ilike(f"%{hotel}%"))
    if date_from: q_cxp = q_cxp.filter(models.Payable.due_date >= date_from)
    if date_to: q_cxp = q_cxp.filter(models.Payable.due_date <= date_to)
    if status_cxp: q_cxp = q_cxp.filter(models.Payable.status == status_cxp)
    cxp_rows = q_cxp.order_by(models.Payable.vendor_name, models.Payable.due_date).all()

    q_others = db.query(models.Other).filter(models.Other.status.notin_(["liquidado","cancelado"]))
    if country: q_others = q_others.filter(models.Other.country == country)
    if date_from: q_others = q_others.filter(models.Other.due_date >= date_from)
    if date_to: q_others = q_others.filter(models.Other.due_date <= date_to)
    others_rows = q_others.order_by(models.Other.due_date).all()

    buf = make_excel(cxc_rows, cxp_rows, others_rows=others_rows, fx=fx_rate)
    filename = f"Dreamart_CXC_CXP_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/others")
def export_others(
    fx_rate: float = Query(default=FX),
    direction: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    # Standalone "Others" export — no CXC/CXP involved.
    # Note: models.Other has no `country` column (unlike Receivable/Payable),
    # so it's intentionally not filterable here.
    q = db.query(models.Other)
    if direction: q = q.filter(models.Other.direction == direction)
    if date_from: q = q.filter(models.Other.due_date >= date_from)
    if date_to: q = q.filter(models.Other.due_date <= date_to)
    if status:
        q = q.filter(models.Other.status == status)
    else:
        q = q.filter(models.Other.status.notin_(["liquidado", "cancelado"]))
    others_rows = q.order_by(models.Other.due_date).all()

    buf = make_others_excel(others_rows, fx=fx_rate)
    filename = f"Dreamart_Otros_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )